#!/usr/bin/env python3
"""
Compiles per-suite pytest-json-report / k6 JSON outputs into:
  1. master_report.xlsx  — full metric + per-suite breakdown workbook
  2. dashboard.html       — dark-mode HTML dashboard, deployable to GitHub Pages

Usage:
  python generate_master_report.py --input-dir downloaded-reports \
      --output-dir build/reports --commit-sha <sha> --run-number <n>
"""
import argparse
import json
import os
from pathlib import Path
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SUITES = [
    ("unit-test-report", "unit_report.json", "🧪 Unit Tests — API", "pytest"),
    ("validation-test-report", "validation_report.json", "✅ Validation Tests", "pytest"),
    ("selenium-web-report", "selenium_report.json", "🌐 Selenium — Website Tests", "pytest"),
    ("appium-android-report", "appium_report.json", "📱 Appium — Android Tests", "pytest"),
    ("load-test-report", "load_report.json", "⚡ Load Testing — Performance", "k6"),
    ("deployment-test-report", "deployment_report.json", "🚀 Deployment Status", "pytest"),
]


def parse_pytest_json(path):
    """Return (passed, failed, skipped, total, duration_s) from a pytest-json-report file."""
    if not path.exists():
        return 0, 0, 0, 0, None
    data = json.loads(path.read_text())
    summary = data.get("summary", {})
    passed = summary.get("passed", 0)
    failed = summary.get("failed", 0)
    skipped = summary.get("skipped", 0)
    total = summary.get("total", passed + failed + skipped)
    duration = data.get("duration")
    return passed, failed, skipped, total, duration


def parse_k6_json(path):
    """Return (passed, failed, skipped, total, duration_s) from a k6 --summary-export file."""
    if not path.exists():
        return 0, 0, 0, 0, None
    data = json.loads(path.read_text())
    metrics = data.get("metrics", {})
    checks = metrics.get("checks", {})
    passes = checks.get("passes", 0)
    fails = checks.get("fails", 0)
    total = passes + fails
    duration = None
    return passes, fails, 0, total, duration


def collect_results(input_dir: Path):
    results = []
    for artifact_dir, filename, label, kind in SUITES:
        report_path = input_dir / artifact_dir / filename
        if kind == "k6":
            passed, failed, skipped, total, duration = parse_k6_json(report_path)
        else:
            passed, failed, skipped, total, duration = parse_pytest_json(report_path)
        status = "PASS" if failed == 0 and total > 0 else ("PASS" if total == 0 else "FAIL")
        results.append({
            "label": label,
            "passed": passed,
            "failed": failed,
            "skipped": skipped,
            "total": total,
            "duration": duration,
            "status": status,
        })
    return results


def build_excel(results, out_path: Path, commit_sha, run_number):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    pass_font = Font(color="15803D", bold=True)
    fail_font = Font(color="B91C1C", bold=True)

    ws["A1"] = "Overall Status"
    total_tests = sum(r["total"] for r in results)
    total_failed = sum(r["failed"] for r in results)
    overall = "PASSED" if total_failed == 0 else "FAILED"
    ws["B1"] = overall
    ws["B1"].font = pass_font if overall == "PASSED" else fail_font

    meta_rows = [
        ("Total Tests Executed", total_tests),
        ("Failed", total_failed),
        ("Pass Rate", f"{((total_tests - total_failed) / total_tests * 100):.1f}%" if total_tests else "N/A"),
        ("Commit SHA", commit_sha),
        ("CI Run Number", f"#{run_number}"),
        ("Generated At (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for i, (k, v) in enumerate(meta_rows, start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)

    start_row = len(meta_rows) + 4
    headers = ["Suite", "Status", "Passed", "Failed", "Skipped", "Total"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    for i, r in enumerate(results, start=start_row + 1):
        ws.cell(row=i, column=1, value=r["label"])
        status_cell = ws.cell(row=i, column=2, value=r["status"])
        status_cell.font = pass_font if r["status"] == "PASS" else fail_font
        ws.cell(row=i, column=3, value=r["passed"])
        ws.cell(row=i, column=4, value=r["failed"])
        ws.cell(row=i, column=5, value=r["skipped"])
        ws.cell(row=i, column=6, value=r["total"])

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 22

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def build_html(results, out_path: Path, commit_sha, run_number):
    total_tests = sum(r["total"] for r in results)
    total_failed = sum(r["failed"] for r in results)
    total_passed = sum(r["passed"] for r in results)
    overall = "PASSED" if total_failed == 0 else "FAILED"
    pass_rate = f"{((total_tests - total_failed) / total_tests * 100):.1f}%" if total_tests else "N/A"

    rows_html = "\n".join(f"""
        <tr>
          <td>{r['label']}</td>
          <td><span class="badge {'pass' if r['status']=='PASS' else 'fail'}">{r['status']}</span></td>
          <td>{r['passed']}</td>
          <td>{r['failed']}</td>
          <td>{r['skipped']}</td>
          <td>{r['total']}</td>
        </tr>""" for r in results)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>OCT Segmentation — Master Test Summary</title>
<style>
  :root {{
    --bg: #0d1117; --panel: #161b22; --border: #30363d;
    --text: #e6edf3; --muted: #8b949e; --accent: #58a6ff;
    --pass: #3fb950; --fail: #f85149;
  }}
  body {{ background: var(--bg); color: var(--text); font-family: -apple-system, Segoe UI, Roboto, sans-serif; margin: 0; padding: 2.5rem; }}
  h1 {{ font-size: 1.6rem; margin-bottom: .25rem; }}
  .sub {{ color: var(--muted); margin-bottom: 2rem; }}
  .grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 1rem; margin-bottom: 2rem; }}
  .card {{ background: var(--panel); border: 1px solid var(--border); border-radius: 10px; padding: 1.1rem 1.3rem; }}
  .card .label {{ color: var(--muted); font-size: .8rem; text-transform: uppercase; letter-spacing: .04em; }}
  .card .value {{ font-size: 1.7rem; font-weight: 700; margin-top: .3rem; }}
  table {{ width: 100%; border-collapse: collapse; background: var(--panel); border: 1px solid var(--border); border-radius: 10px; overflow: hidden; }}
  th, td {{ padding: .8rem 1rem; text-align: left; border-bottom: 1px solid var(--border); }}
  th {{ background: #1c2128; color: var(--muted); font-size: .8rem; text-transform: uppercase; }}
  tr:last-child td {{ border-bottom: none; }}
  .badge {{ padding: .2rem .6rem; border-radius: 999px; font-size: .78rem; font-weight: 600; }}
  .badge.pass {{ background: rgba(63,185,80,.15); color: var(--pass); }}
  .badge.fail {{ background: rgba(248,81,73,.15); color: var(--fail); }}
  footer {{ margin-top: 2rem; color: var(--muted); font-size: .8rem; font-style: italic; }}
</style>
</head>
<body>
  <h1>Automated Retinal Layer Segmentation — Master Test Summary</h1>
  <div class="sub">Commit {commit_sha[:7] if commit_sha else 'N/A'} · CI Run #{run_number} · Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}</div>

  <div class="grid">
    <div class="card"><div class="label">Overall Status</div><div class="value" style="color: var(--{'pass' if overall=='PASSED' else 'fail'})">{overall}</div></div>
    <div class="card"><div class="label">Total Tests</div><div class="value">{total_tests}</div></div>
    <div class="card"><div class="label">Passed</div><div class="value" style="color: var(--pass)">{total_passed}</div></div>
    <div class="card"><div class="label">Failed</div><div class="value" style="color: var(--fail)">{total_failed}</div></div>
    <div class="card"><div class="label">Pass Rate</div><div class="value">{pass_rate}</div></div>
  </div>

  <table>
    <thead>
      <tr><th>Suite</th><th>Status</th><th>Passed</th><th>Failed</th><th>Skipped</th><th>Total</th></tr>
    </thead>
    <tbody>
      {rows_html}
    </tbody>
  </table>

  <footer>Report generated by OCT Segmentation Master CI/CD Pipeline</footer>
</body>
</html>"""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(html)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input-dir", required=True)
    ap.add_argument("--output-dir", required=True)
    ap.add_argument("--commit-sha", default="unknown")
    ap.add_argument("--run-number", default="0")
    args = ap.parse_args()

    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)

    results = collect_results(input_dir)
    build_excel(results, output_dir / "master_report.xlsx", args.commit_sha, args.run_number)
    build_html(results, output_dir / "dashboard.html", args.commit_sha, args.run_number)
    # GitHub Pages needs an index.html
    (output_dir / "index.html").write_text((output_dir / "dashboard.html").read_text())

    print(f"Wrote reports to {output_dir}")


if __name__ == "__main__":
    main()
